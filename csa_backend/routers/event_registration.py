import logging
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from supabase import create_client, Client
import os
from typing import Optional
from io import BytesIO
from services.auth_services import verify_admin_token
from services.event_email_scheduler import process_reminder_emails_for_tomorrow, process_thank_you_emails
from services.event_email_service import send_confirmation_email
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Initialize router
event_registration_router = APIRouter()

# Supabase setup - will be created in functions
def get_supabase_client():
    """Get Supabase client with environment variables"""
    supabase_url = os.getenv("CSA_SUPABASE_URL")
    supabase_key = os.getenv("CSA_SUPABASE_SERVICE_KEY")
    if not supabase_url or not supabase_key:
        raise Exception("Supabase credentials not found")
    return create_client(supabase_url, supabase_key)

# Event Registration Models
class EventRegistrationRequest(BaseModel):
    user_id: str
    event_id: str

class EventRegistrationResponse(BaseModel):
    id: str
    user_id: str
    event_id: str
    message: str

@event_registration_router.post("/event-registrations", response_model=EventRegistrationResponse)
async def create_event_registration(registration: EventRegistrationRequest):
    """
    Register a user for an event.
    Creates a new event registration record in the database.
    """
    logging.info(f"Attempting to register user {registration.user_id} for event {registration.event_id}")
    
    try:
        supabase = get_supabase_client()
        
        # Validate user exists in either users or admins table
        user_response = supabase.table("users").select("id, email, name").eq("id", registration.user_id).limit(1).execute()
        admin_response = supabase.table("admins").select("id, email, name").eq("id", registration.user_id).limit(1).execute()
        
        if not user_response.data and not admin_response.data:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Get user details for email
        user_data = user_response.data[0] if user_response.data else admin_response.data[0]
        user_email = user_data.get("email")
        user_name = user_data.get("name") or "Valued Member"
        
        # Validate event exists and get full details for email
        event_response = supabase.table("events").select("id, title, capacity, attendees, date_time, location, slug").eq("id", registration.event_id).limit(1).execute()
        if not event_response.data:
            raise HTTPException(status_code=404, detail="Event not found")
        
        event_data = event_response.data[0]
        
        # Check if user is already registered
        existing_registration = supabase.table("event_registrations").select("id").eq(
            "user_id", registration.user_id
        ).eq("event_id", registration.event_id).limit(1).execute()
        
        if existing_registration.data:
            raise HTTPException(status_code=400, detail="User already registered for this event")
        
        # Check if event is at capacity
        if event_data["attendees"] >= event_data["capacity"]:
            raise HTTPException(status_code=400, detail="Event is at full capacity")
        
        # Create new registration with pending status (will update after email sent)
        registration_response = supabase.table("event_registrations").insert({
            "user_id": registration.user_id,
            "event_id": registration.event_id,
            "email_status": "pending"
        }).execute()
        
        if not registration_response.data:
            raise HTTPException(status_code=500, detail="Failed to create registration")
        
        # Update event attendees count
        new_attendees = event_data["attendees"] + 1
        supabase.table("events").update({
            "attendees": new_attendees
        }).eq("id", registration.event_id).execute()
        
        registration_id = registration_response.data[0]["id"]
        
        # Send confirmation email immediately
        if user_email:
            try:
                event_title = event_data.get("title", "Event")
                event_date_time = event_data.get("date_time", "")
                event_location = event_data.get("location", "")
                event_slug = event_data.get("slug")
                
                email_sent = await send_confirmation_email(
                    to_email=user_email,
                    user_name=user_name,
                    event_title=event_title,
                    event_date_time=event_date_time,
                    event_location=event_location,
                    event_slug=event_slug,
                )
                
                if email_sent:
                    # Update registration with confirmation timestamp
                    supabase.table("event_registrations").update({
                        "confirmation_sent_at": datetime.utcnow().isoformat(),
                        "email_status": "confirmation_sent"
                    }).eq("id", registration_id).execute()
                    logging.info(f"Confirmation email sent to {user_email} for event {event_title}")
                else:
                    # Update to failed status if email didn't send
                    supabase.table("event_registrations").update({
                        "email_status": "failed",
                        "email_error": "Failed to send confirmation email"
                    }).eq("id", registration_id).execute()
                    logging.warning(f"Failed to send confirmation email to {user_email}, but registration was created")
            except Exception as e:
                logging.error(f"Error sending confirmation email: {e}")
                # Update to failed status
                supabase.table("event_registrations").update({
                    "email_status": "failed",
                    "email_error": str(e)
                }).eq("id", registration_id).execute()
                # Don't fail the registration if email fails
        else:
            logging.warning(f"User {registration.user_id} has no email address, skipping confirmation email")
            # Update status to indicate no email
            supabase.table("event_registrations").update({
                "email_status": "no_email"
            }).eq("id", registration_id).execute()
        
        logging.info(f"Registration created: {registration_id}")
        return EventRegistrationResponse(
            id=registration_id,
            user_id=registration.user_id,
            event_id=registration.event_id,
            message="Registration successful"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating registration: {e}")
        raise HTTPException(status_code=500, detail=f"Registration failed: {str(e)}")

@event_registration_router.get("/event-registrations/{user_id}")
async def get_user_registrations(user_id: str):
    """
    Get all event registrations for a specific user.
    """
    logging.info(f"Fetching registrations for user {user_id}")
    
    try:
        supabase = get_supabase_client()
        
        # Validate user exists in either users or admins table
        user_response = supabase.table("users").select("id").eq("id", user_id).limit(1).execute()
        admin_response = supabase.table("admins").select("id").eq("id", user_id).limit(1).execute()
        
        if not user_response.data and not admin_response.data:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Get user's registrations with event details
        registrations_response = supabase.table("event_registrations").select(
            "id, event_id, events(title, date_time, location, slug)"
        ).eq("user_id", user_id).execute()
        
        return {
            "registrations": registrations_response.data if registrations_response.data else []
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching registrations: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch registrations: {str(e)}")

@event_registration_router.delete("/event-registrations/{registration_id}")
async def cancel_event_registration(registration_id: str):
    """
    Cancel an event registration.
    """
    logging.info(f"Cancelling registration {registration_id}")
    
    try:
        supabase = get_supabase_client()
        
        # Get registration details
        registration_response = supabase.table("event_registrations").select(
            "id, user_id, event_id, events(attendees)"
        ).eq("id", registration_id).limit(1).execute()
        
        if not registration_response.data:
            raise HTTPException(status_code=404, detail="Registration not found")
        
        registration_data = registration_response.data[0]
        event_id = registration_data["event_id"]
        current_attendees = registration_data["events"]["attendees"]
        
        # Delete registration
        delete_response = supabase.table("event_registrations").delete().eq("id", registration_id).execute()
        
        if not delete_response.data:
            raise HTTPException(status_code=500, detail="Failed to cancel registration")
        
        # Update event attendees count
        new_attendees = max(0, current_attendees - 1)
        supabase.table("events").update({
            "attendees": new_attendees
        }).eq("id", event_id).execute()
        
        logging.info(f"Registration {registration_id} cancelled")
        return {"message": "Registration cancelled successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error cancelling registration: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to cancel registration: {str(e)}")

@event_registration_router.post("/simple-registration")
async def simple_event_registration(registration: EventRegistrationRequest):
    """
    Register a user for an event without authentication.
    Sends confirmation email immediately.
    Reminder and thank-you emails are sent by scheduled jobs based on dates.
    """
    logging.info(f"Registering user {registration.user_id} for event {registration.event_id}")
    
    try:
        supabase = get_supabase_client()
        
        # Get user details for email
        user_response = supabase.table("users").select("id, email, name").eq("id", registration.user_id).limit(1).execute()
        admin_response = supabase.table("admins").select("id, email, name").eq("id", registration.user_id).limit(1).execute()
        
        if not user_response.data and not admin_response.data:
            raise HTTPException(status_code=404, detail="User not found")
        
        user_data = user_response.data[0] if user_response.data else admin_response.data[0]
        user_email = user_data.get("email")
        user_name = user_data.get("name") or "Valued Member"
        
        # Get event details (need full details for email)
        event_response = supabase.table("events").select("id, title, date_time, location, slug, attendees, capacity").eq("id", registration.event_id).limit(1).execute()
        if not event_response.data:
            raise HTTPException(status_code=404, detail="Event not found")
        
        event_data = event_response.data[0]
        
        # Check if user is already registered
        existing_registration = supabase.table("event_registrations").select("id").eq(
            "user_id", registration.user_id
        ).eq("event_id", registration.event_id).limit(1).execute()
        
        if existing_registration.data:
            raise HTTPException(status_code=400, detail="User already registered for this event")
        
        # Check if event is at capacity
        if event_data["attendees"] >= event_data["capacity"]:
            raise HTTPException(status_code=400, detail="Event is at full capacity")
        
        # Create new registration with pending status (will update after email sent)
        registration_response = supabase.table("event_registrations").insert({
            "user_id": registration.user_id,
            "event_id": registration.event_id,
            "email_status": "pending"
        }).execute()
        
        if not registration_response.data:
            raise HTTPException(status_code=500, detail="Failed to create registration")
        
        # Update event attendees count
        new_attendees = event_data["attendees"] + 1
        supabase.table("events").update({
            "attendees": new_attendees
        }).eq("id", registration.event_id).execute()
        
        registration_id = registration_response.data[0]["id"]
        
        # Send confirmation email immediately
        if user_email:
            try:
                event_title = event_data.get("title", "Event")
                event_date_time = event_data.get("date_time", "")
                event_location = event_data.get("location", "")
                event_slug = event_data.get("slug")
                
                email_sent = await send_confirmation_email(
                    to_email=user_email,
                    user_name=user_name,
                    event_title=event_title,
                    event_date_time=event_date_time,
                    event_location=event_location,
                    event_slug=event_slug,
                )
                
                if email_sent:
                    # Update registration with confirmation timestamp
                    supabase.table("event_registrations").update({
                        "confirmation_sent_at": datetime.utcnow().isoformat(),
                        "email_status": "confirmation_sent"
                    }).eq("id", registration_id).execute()
                    logging.info(f"Confirmation email sent to {user_email} for event {event_title}")
                else:
                    # Update to failed status if email didn't send
                    supabase.table("event_registrations").update({
                        "email_status": "failed",
                        "email_error": "Failed to send confirmation email"
                    }).eq("id", registration_id).execute()
                    logging.warning(f"Failed to send confirmation email to {user_email}, but registration was created")
            except Exception as e:
                logging.error(f"Error sending confirmation email: {e}")
                # Update to failed status
                supabase.table("event_registrations").update({
                    "email_status": "failed",
                    "email_error": str(e)
                }).eq("id", registration_id).execute()
                # Don't fail the registration if email fails
        else:
            logging.warning(f"User {registration.user_id} has no email address, skipping confirmation email")
            # Update status to indicate no email
            supabase.table("event_registrations").update({
                "email_status": "no_email"
            }).eq("id", registration_id).execute()
        
        logging.info(f"Registration created: {registration_id}")
        return {
            "id": registration_id,
            "user_id": registration.user_id,
            "event_id": registration.event_id,
            "message": "Registration successful"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in registration: {e}")
        raise HTTPException(status_code=500, detail=f"Registration failed: {str(e)}")

@event_registration_router.get("/event-attendees/{event_id}")
async def get_event_attendees(event_id: str):
    """
    Get the current attendees count for a specific event.
    """
    try:
        supabase = get_supabase_client()
        
        # Get event details
        event_response = supabase.table("events").select("id, attendees, capacity").eq("id", event_id).execute()
        if not event_response.data:
            raise HTTPException(status_code=404, detail="Event not found")
        
        event = event_response.data[0]
        
        # Return the attendees count from events table (which is kept up-to-date by registrations)
        return {
            "event_id": event_id,
            "attendees": event["attendees"],
            "capacity": event["capacity"],
            "spots_left": event["capacity"] - event["attendees"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting event attendees: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting attendees: {e}")

@event_registration_router.get("/event-registered-users/{event_id}")
async def get_event_registered_users(event_id: str):
    """
    Get all registered users for a specific event with their details.
    """
    try:
        supabase = get_supabase_client()
        
        # Get event details to verify it exists
        event_response = supabase.table("events").select("id, title").eq("id", event_id).execute()
        if not event_response.data:
            raise HTTPException(status_code=404, detail="Event not found")
        
        # Get all registrations for this event
        registrations_response = supabase.table("event_registrations").select("user_id, updated_at").eq("event_id", event_id).execute()
        
        if not registrations_response.data:
            return {
                "event_id": event_id,
                "registered_users": [],
                "count": 0
            }
        
        # Get user details for each registration
        user_ids = [reg["user_id"] for reg in registrations_response.data]
        
        # Fetch users from users table
        users_response = supabase.table("users").select("id, name, email, company_name, role, avatar_url").in_("id", user_ids).execute()
        users = users_response.data if users_response.data else []
        
        # Also check admins table in case any admins registered
        admins_response = supabase.table("admins").select("id, name, email").in_("id", user_ids).execute()
        admins = admins_response.data if admins_response.data else []
        
        # Combine users and admins, adding a type field
        all_users = []
        for user in users:
            user["user_type"] = "user"
            all_users.append(user)
        
        for admin in admins:
            admin["user_type"] = "admin"
            admin["company_name"] = "CSA Admin"
            admin["role"] = "Administrator"
            admin["avatar_url"] = None
            all_users.append(admin)
        
        # Create a map of user details by ID
        user_map = {user["id"]: user for user in all_users}
        
        # Combine registration data with user details
        registered_users = []
        for reg in registrations_response.data:
            user_id = reg["user_id"]
            if user_id in user_map:
                user_data = user_map[user_id].copy()
                user_data["registered_at"] = reg.get("updated_at", "")
                registered_users.append(user_data)
        
        return {
            "event_id": event_id,
            "registered_users": registered_users,
            "count": len(registered_users)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting registered users for event: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting registered users: {e}")

@event_registration_router.delete("/event-registrations/delete/{event_id}/{user_id}")
async def delete_event_registration(event_id: str, user_id: str, token_data: dict = Depends(verify_admin_token)):
    """
    Delete a user's registration for a specific event (Admin only).
    Also updates the event's attendee count.
    """
    admin_email = token_data.get("email", "Unknown")
    logging.info(f"Admin {admin_email} is deleting registration for user {user_id} from event {event_id}")
    
    try:
        supabase = get_supabase_client()
        
        # Check if registration exists
        reg_check = supabase.table("event_registrations").select("id").eq("user_id", user_id).eq("event_id", event_id).execute()
        
        if not reg_check.data or len(reg_check.data) == 0:
            raise HTTPException(status_code=404, detail="Registration not found")
        
        # Get current event attendees count
        event_resp = supabase.table("events").select("id, attendees").eq("id", event_id).limit(1).execute()
        if not event_resp.data:
            raise HTTPException(status_code=404, detail="Event not found")
        
        current_attendees = event_resp.data[0].get("attendees", 0)
        
        # Delete the registration
        delete_resp = supabase.table("event_registrations").delete().eq("user_id", user_id).eq("event_id", event_id).execute()
        
        if not delete_resp.data:
            raise HTTPException(status_code=500, detail="Failed to delete registration")
        
        # Update event attendees count
        new_attendees = max(0, current_attendees - 1)
        supabase.table("events").update({"attendees": new_attendees}).eq("id", event_id).execute()
        
        logging.info(f"Admin {admin_email} successfully deleted registration and updated attendees: {current_attendees} -> {new_attendees}")
        
        return {
            "message": "Registration deleted successfully",
            "event_id": event_id,
            "user_id": user_id,
            "previous_attendees": current_attendees,
            "updated_attendees": new_attendees
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting event registration: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting registration: {str(e)}")

@event_registration_router.get("/debug/event-registrations")
async def debug_event_registrations():
    """
    Debug endpoint to check event registration setup
    """
    try:
        supabase = get_supabase_client()
        
        # Check if event_registrations table exists
        table_check = supabase.table("event_registrations").select("id").limit(1).execute()
        
        # Check users count
        users_count = supabase.table("users").select("id", count="exact").execute()
        
        # Check events count
        events_count = supabase.table("events").select("id", count="exact").execute()
        
        return {
            "status": "ok",
            "event_registrations_table": "exists" if table_check.data is not None else "missing",
            "users_count": users_count.count if users_count.count else 0,
            "events_count": events_count.count if events_count.count else 0,
            "supabase_connected": True
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "supabase_connected": False
        }

@event_registration_router.post("/event-emails/process")
async def process_event_emails():
    """
    Endpoint for Supabase cron job to trigger event email processing.
    Called by the scheduled cron job at 8am PST daily.
    
    Processes both:
    - Reminder emails (for events happening tomorrow)
    - Thank-you emails (for events that completed yesterday)
    """
    try:
        logging.info("Processing reminder emails for events happening tomorrow...")
        reminder_count = await process_reminder_emails_for_tomorrow()
        logging.info(f"Reminder email processing completed. Sent {reminder_count} reminder(s).")
        
        logging.info("Processing thank-you emails for events that completed yesterday...")
        thank_you_count = await process_thank_you_emails()
        logging.info(f"Thank-you email processing completed. Sent {thank_you_count} thank-you email(s).")
        
        return {
            "success": True,
            "reminder_emails_sent": reminder_count,
            "thank_you_emails_sent": thank_you_count,
            "message": f"Processed emails. Sent {reminder_count} reminder(s) and {thank_you_count} thank-you email(s)."
        }
    except Exception as e:
        logging.error(f"Error processing event emails: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error processing event emails: {str(e)}"
        )

@event_registration_router.get("/export-attendees/{event_id}")
async def export_event_attendees_to_excel(event_id: str, token_data: dict = Depends(verify_admin_token)):
    """
    Export all attendees for a specific event to an Excel file.
    Admin only endpoint.
    """
    try:
        supabase = get_supabase_client()
        
        # Get event details
        event_response = supabase.table("events").select("id, title, date_time, location").eq("id", event_id).execute()
        if not event_response.data:
            raise HTTPException(status_code=404, detail="Event not found")
        
        event = event_response.data[0]
        
        # Get all registrations for this event
        registrations_response = supabase.table("event_registrations").select("event_id, user_id, updated_at").eq("event_id", event_id).execute()
        registrations = registrations_response.data if registrations_response.data else []
        
        if not registrations:
            # No registrations, return empty Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendees"
            ws.append(["No attendees found for this event"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"attendees_{event.get('title', 'event').replace(' ', '_')}_{timestamp}.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        # Get all unique user IDs
        user_ids = list(set([reg["user_id"] for reg in registrations]))
        
        # Get user details
        users_response = supabase.table("users").select("id, name, email, company_name, role, avatar_url").in_("id", user_ids).execute()
        users = users_response.data if users_response.data else []
        
        # Get admin details (in case admins registered)
        admins_response = supabase.table("admins").select("id, name, email").in_("id", user_ids).execute()
        admins = admins_response.data if admins_response.data else []
        
        # Create user map
        user_map = {}
        for user in users:
            user_map[user["id"]] = {
                **user,
                "user_type": "user"
            }
        for admin in admins:
            user_map[admin["id"]] = {
                **admin,
                "user_type": "admin",
                "company_name": "CSA Admin",
                "role": "Administrator",
                "avatar_url": None
            }
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendees"
        
        # Add event info header
        ws.append(["Event Information"])
        ws.append(["Event Title", event.get("title", "")])
        ws.append(["Event Date", event.get("date_time", "")])
        ws.append(["Event Location", event.get("location", "")])
        ws.append([])  # Empty row
        
        # Header row
        headers = [
            "Name", "Email", "Company", "Role", "User Type",
            "Registration Date"
        ]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[6]:  # Row 6 is the header row (after event info)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for reg in registrations:
            user_id = reg["user_id"]
            user = user_map.get(user_id, {})
            
            # Format registration date (using updated_at as registration timestamp)
            reg_date = ""
            if reg.get("updated_at"):
                try:
                    dt_str = reg["updated_at"]
                    if dt_str.endswith("Z"):
                        dt_str = dt_str[:-1] + "+00:00"
                    dt = datetime.fromisoformat(dt_str)
                    reg_date = dt.strftime("%Y-%m-%d %H:%M")
                except Exception as e:
                    logging.debug(f"Error parsing registration date: {e}")
                    reg_date = reg.get("updated_at", "")
            
            row = [
                user.get("name", "Unknown"),
                user.get("email", ""),
                user.get("company_name", ""),
                user.get("role", ""),
                user.get("user_type", "user"),
                reg_date
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create BytesIO buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with event title and timestamp
        event_title_safe = "".join(c for c in event.get("title", "event") if c.isalnum() or c in (' ', '-', '_')).strip()[:50]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendees_{event_title_safe.replace(' ', '_')}_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error exporting event attendees to Excel: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error exporting attendees: {str(e)}"
        )
